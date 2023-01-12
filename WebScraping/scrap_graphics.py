import requests
from selenium import webdriver
from bs4 import BeautifulSoup
from xml.etree import ElementTree
import csv
import xlsxwriter
import openpyxl
import time
import os
import json
import datetime



def scrap_graphics(url,champ,number_of_champ):
    driver = webdriver.Chrome(executable_path = 'C:/WebDriver/bin/chromedriver.exe')
    print('fortnite',champ)
    url_esp = url+champ.replace(' ','').replace("'",'').replace(".",'').replace("wukong",'monkeyking').replace("glasc","")
    print('url',url_esp)
    driver.get(url_esp)
    time.sleep(5)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    scripts = soup.findAll('script')

    #Excel Pop
    workbookPopHistory = openpyxl.load_workbook(filename = r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoPopComplete.xlsx')
    worksheetPopHistory = workbookPopHistory.active

    #Excel WR
    workbookWRHistory = openpyxl.load_workbook(filename = r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoWRComplete.xlsx')
    worksheetWRHistory = workbookWRHistory.active

    #Excel BR
    workbookBRHistory  = openpyxl.load_workbook(filename = r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoBRComplete.xlsx')
    worksheetBRHistory = workbookBRHistory.active



    wrhistoryscript = ''

    for script in scripts:
        if script.text.find('graphFuncgraphDD5')!=-1:
            popularityhistoryscript = script.text
            data = popularityhistoryscript.split('data: ')
            lines = data[1].split('lines')[0]
            ph = lines.split(',\n')[0]
            popularityhistoryall = json.loads(ph)
            popularityhistory = [row[1] for row in popularityhistoryall]
            popularityhistorydate = [datetime.datetime.fromtimestamp(row[0]/1000).strftime('%Y-%m-%d') for row in popularityhistoryall]
            popularityhistorydate = popularityhistorydate[::-1]
            popularityhistory = popularityhistory[::-1]
            # popularityhistorydate = [row[0] for row in popularityhistoryall]
        if script.text.find('graphFuncgraphDD6')!=-1:
            wrhistoryscript = script.text
            data = wrhistoryscript.split('data: ')
            lines = data[1].split('lines')[0]
            wh = lines.split(',\n')[0]
            wrhistoryall = json.loads(wh)
            wrhistory = [row[1] for row in wrhistoryall]
            #wrhistorydate = [datetime.datetime.fromtimestamp(row[0]/1000).strftime('%Y-%m-%d') for row in wrhistoryall]
            wrhistorydate = [row[0] for row in wrhistoryall]
            wrhistorydate = wrhistorydate[::-1]
        if script.text.find('graphFuncgraphDD7')!=-1:
            brhistoryscript = script.text
            data = brhistoryscript.split('data: ')
            lines = data[1].split('lines')[0]
            bh = lines.split(',\n')[0]
            brhistoryall = json.loads(bh)
            brhistory = [row[1] for row in brhistoryall]
            #brhistorydate = [datetime.datetime.fromtimestamp(row[0]/1000).strftime('%Y-%m-%d') for row in brhistoryall]
            brhistorydate = [row[0] for row in brhistoryall]
            brhistorydate = brhistorydate[::-1]

    for i in range(len(popularityhistorydate)):
        worksheetPopHistory.cell(row=i+1,column=1).value = popularityhistorydate[i-1]
        worksheetPopHistory.cell(row=i+1,column=1+number_of_champ).value = popularityhistory[i-1]
        
        worksheetWRHistory.cell(row=i+1,column=1).value = popularityhistorydate[i-1]
        worksheetWRHistory.cell(row=i+1,column=1+number_of_champ).value = wrhistory[i-1]
       
        worksheetBRHistory.cell(row=i+1,column=1).value = popularityhistorydate[i-1]
        worksheetBRHistory.cell(row=i+1,column=1+number_of_champ).value = brhistory[i-1]

    worksheetPopHistory.cell(row=1,column=1).value = "date"
    worksheetWRHistory.cell(row=1,column=1).value = "date"
    worksheetBRHistory.cell(row=1,column=1).value = "date"

    worksheetPopHistory.cell(row=1,column=1+number_of_champ).value = champ
    worksheetWRHistory.cell(row=1,column=1+number_of_champ).value = champ
    worksheetBRHistory.cell(row=1,column=1+number_of_champ).value = champ
    
    workbookPopHistory.save(r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoPopComplete.xlsx')
    workbookWRHistory.save(r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoWRComplete.xlsx')
    workbookBRHistory.save(r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoBRComplete.xlsx')

    driver.close()