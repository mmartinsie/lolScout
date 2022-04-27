import requests
from selenium import webdriver
from bs4 import BeautifulSoup
from xml.etree import ElementTree
import csv
import xlsxwriter
import openpyxl
import time
import os
import json
import datetime



def scrap_champ(url,champ,number_of_champ):
    driver = webdriver.Chrome(executable_path = 'C:/WebDriver/bin/chromedriver.exe')
    print('fortnite',champ)
    url_esp = url+champ.replace(' ','').replace("'",'').replace(".",'').replace("wukong",'monkeyking').replace("glasc","")
    print('url',url_esp)
    driver.get(url_esp)
    time.sleep(5)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    scripts = soup.findAll('script')
       
    #with open('champs_information.csv', 'a', newline='') as csv_file:
        #headers = ['Name', 'Popularity', 'WR','Banrate', 'Main', 'Pentakills', 'Gold', 'Minions', 'Wards', 'Damage', 'Phistory', 'PhistoryDate', 'WRhistory', 'WRhistoryDate', 'BRhistory', 'BRhistoryDate'] 
        #writer = csv.DictWriter(csv_file, fieldnames=headers)

    # workbook = xlsxwriter.Workbook(r'C:\Users\Manuel Martín Sierra\Documents\lolScout\WebScraping\champInfoVersion.xlsx')
    workbook = openpyxl.load_workbook(filename = r'C:\Users\Manuel Martín Sierra\Documents\lolScout\WebScraping\champInfoVersion.xlsx')
    worksheet =  workbook.active

    workbookHistory  = openpyxl.Workbook()
    workbookHistory.create_sheet(title="V12.6")
    worksheetHistory = workbookHistory.active


    wrhistoryscript = ''

    for script in scripts:
        if script.text.find('graphFuncgraphDD5')!=-1:
            popularityhistoryscript = script.text
            data = popularityhistoryscript.split('data: ')
            lines = data[1].split('lines')[0]
            ph = lines.split(',\n')[0]
            popularityhistoryall = json.loads(ph)
            popularityhistory = [row[1] for row in popularityhistoryall]
            popularityhistorydate = [datetime.datetime.fromtimestamp(row[0]/1000).strftime('%Y-%m-%d') for row in popularityhistoryall]
            # popularityhistorydate = [row[0] for row in popularityhistoryall]
        if script.text.find('graphFuncgraphDD6')!=-1:
            wrhistoryscript = script.text
            data = wrhistoryscript.split('data: ')
            lines = data[1].split('lines')[0]
            wh = lines.split(',\n')[0]
            wrhistoryall = json.loads(wh)
            wrhistory = [row[1] for row in wrhistoryall]
            #wrhistorydate = [datetime.datetime.fromtimestamp(row[0]/1000).strftime('%Y-%m-%d') for row in wrhistoryall]
            wrhistorydate = [row[0] for row in wrhistoryall]
        if script.text.find('graphFuncgraphDD7')!=-1:
            brhistoryscript = script.text
            data = wrhistoryscript.split('data: ')
            lines = data[1].split('lines')[0]
            bh = lines.split(',\n')[0]
            brhistoryall = json.loads(bh)
            brhistory = [row[1] for row in brhistoryall]
            #brhistorydate = [datetime.datetime.fromtimestamp(row[0]/1000).strftime('%Y-%m-%d') for row in brhistoryall]
            brhistorydate = [row[0] for row in brhistoryall]

    popularity =  driver.find_element_by_xpath('//*[@id="graphDD1"]').text
    a = slice(len(popularity)-1)
    popularity = popularity[a]
    print(popularity)
    wr =  driver.find_element_by_xpath('//*[@id="graphDD2"]').text
    a = slice(len(wr)-1)
    wr = wr[a]
    print(wr)
    banrate =  driver.find_element_by_xpath('//*[@id="graphDD3"]').text
    a = slice(len(banrate)-1)
    banrate = banrate[a]
    print(banrate)
    main =  driver.find_element_by_xpath('//*[@id="graphDD4"]').text
    a = slice(len(main)-1)
    main = main[a]
    print(main)
    pentakills =  driver.find_element_by_xpath('//*[@id="mainContent"]/div[2]/div[2]/div[5]/div[1]/div/a/div[1]').text
    print(pentakills)
    gold =  driver.find_element_by_xpath('//*[@id="mainContent"]/div[2]/div[2]/div[6]/div[1]/div/div[1]').text
    print(gold)
    minions =  driver.find_element_by_xpath('//*[@id="mainContent"]/div[2]/div[2]/div[6]/div[2]/div/div[1]').text
    print(minions)
    wards =  driver.find_element_by_xpath('/html/body/div[2]/div[3]/div[3]/div[1]/div[2]/div[2]/div[2]/div[6]/div[3]/div/div[1]').text
    print(wards)
    damage =  driver.find_element_by_xpath('/html/body/div[2]/div[3]/div[3]/div[1]/div[2]/div[2]/div[2]/div[6]/div[4]/div/div[1]').text
    print(damage)

    complete_info =[champ, popularity, wr, banrate, main, pentakills, minions, wards]
    worksheet.append(complete_info)

    print('overwatch',len(popularityhistorydate)-1)
    for i in range(len(popularityhistorydate)-1):
        worksheetHistory.cell(row=i+1,column=1).value = popularityhistorydate[i-1]
        worksheetHistory.cell(row=i+1,column=2).value = wrhistory[i-1]
        worksheetHistory.cell(row=i+1,column=3).value = popularityhistory[i-1]
        worksheetHistory.cell(row=i+1,column=4).value = brhistory[i-1]

    
    workbook.save('champInfoVersion.xlsx')
    workbookHistory.save(r'C:\Users\Manuel Martín Sierra\Documents\lolScout\WebScraping\champsHistory\champInfoHistory'+str(champ)+'.xlsx')
    driver.close()