# Import libraries
from selenium import webdriver
from bs4 import BeautifulSoup
from xml.etree import ElementTree
from scrap_graphics import scrap_graphics
import csv
import time
import os
import json
import xlsxwriter
from openpyxl import Workbook

# Global variables
NUMBER_OF_CHAMPS = 161
# NUMBER_OF_CHAMPS = 5
URL='https://www.leagueofgraphs.com/en/champions/stats/'
URL_ESP='https://www.leagueofgraphs.com/es/champions/stats/'

# Configuration of webdriver to use Google Chrome browser
driver = webdriver.Chrome(executable_path = 'C:/WebDriver/bin/chromedriver.exe')
driver.get(URL)
time.sleep(3)
# driver.maximize_window()
driver.find_element_by_xpath('//*[@id="championsFilter"]/a').click()

allChampionList = driver.find_element_by_xpath('//*[@id="drop-champions"]/ul')
champ_bucket = []

wbPop = Workbook()
wbPop.save(filename = r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoPopComplete.xlsx')

wbWR = Workbook()
wbWR.save(filename = r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoWRComplete.xlsx')

wbBR = Workbook()
wbBR.save(filename = r'C:\Users\Manuel Martín Sierra\Documents\TFG\Series temporales\champsHistoryGraphics\champInfoBRComplete.xlsx')



i = 2
# NUMBER_OF_CHAMPS
for i in range(2, 7):
   
    try:
        # driver.find_element_by_xpath('//*[@id="drop-champions"]/ul/li['+str(i)+']').click()
        champ_bucket.append(driver.find_element_by_xpath('//*[@id="drop-champions"]/ul/li['+str(i)+']').text.lower())
        
    except:
        pass
print(champ_bucket)
print(len(champ_bucket))

number_of_champ=1
for champ_string in champ_bucket:
    scrap_graphics(URL_ESP,champ_string,number_of_champ)
    number_of_champ+=1

    # continue_link = driver.find_element_by_partial_link_text('graphFuncgraphDD5')
    # print(continue_link)
    
# print(allChampionList.text)

# champUrlList = allChampionList.find_all("li")
# print(champUrlList)
driver.close()